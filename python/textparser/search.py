import os
import sys
import re, time
import glob
from docx import Document
from openpyxl import load_workbook
import PyPDF2
import pdfminer.high_level
import shutil
import progressbar
import docx2txt
from progressbar import widgets

i = 0
matches = {}
desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop\\') #путь до рабочего стола
foldername = input('Пожалуйста введите имя папки с файлами на рабочем столе: ') + "\\" 
help_dir_name = desktop +'costs\\'
path = desktop + foldername #путь до папки со всеми файлами для работы скрипта
wb = load_workbook(desktop + 'List.xlsx')
sheet = wb[wb.sheetnames[0]]

errors = open(desktop + 'errors.txt', 'w')

os.mkdir(help_dir_name) #вспомогательная деректория 


def find_key(what, where):
	reg = "\\b"+what+"\\b"
	result = len(re.findall(reg, where, re.IGNORECASE))
	return result

pdf_files = glob.glob(path + '/**/*.pdf', recursive=True)
doc_files = glob.glob(path + '/**/*.doc', recursive=True)
docx_files = glob.glob(path + '/**/*.docx', recursive=True)

max_files = len(pdf_files) + len(doc_files) + len(docx_files)

with progressbar.ProgressBar(max_value=max_files) as bar:

	for path_to_file in pdf_files:
		try:
			file_name = os.path.basename(path_to_file)
			with open(help_dir_name + file_name + '-copy.txt', 'w') as subFile: #костыляем, переписываем пдф в тхт
				pdf = open(path_to_file, 'rb')
				pdfminer.high_level.extract_text_to_fp(pdf, subFile)

			
			for cell in sheet['A']:
				pattern = cell.value.strip()
				ins = 0
				with open(help_dir_name + file_name + '-copy.txt','r') as subFile:
					ins = 0
					for line in subFile:
						ins += find_key(pattern, line)
					if ins > 5:
						if matches.get(pattern) != None:
							matches[pattern] += [path_to_file]
						else:
							matches[pattern] = [path_to_file]
						i += 1
						bar.update(i)
			pdf.close()
		except:

			print('\n Данный файл ' + path_to_file + ' вызвал ошибку', file=errors)


	for path_to_file in docx_files:
		try:
			file_name = os.path.basename(path_to_file)
			
			text = docx2txt.process(path_to_file)
			with open(help_dir_name + file_name + '-copy.txt','w') as subFile:
				print(text, file=subFile)
			
			for cell in sheet['A']:
				pattern = cell.value.strip()
				with open(help_dir_name + file_name + '-copy.txt','r') as subFile:
					ins = 0
					for line in subFile:
						ins += find_key(pattern, line)
					if ins > 5:

						if matches.get(pattern) != None:
							matches[pattern] += [path_to_file]
						else:
							matches[pattern] = [path_to_file]
						
						i += 1
						bar.update(i)
		except:

			print('\n Данный файл ' + path_to_file + ' вызвал ошибку', file=errors)


	for path_to_file in doc_files:
		try:
			file_name = os.path.basename(path_to_file)
			sys_antiword = r'C:\antiword\antiword.exe -m cp1251.txt -w 0 '
			sys_query = sys_antiword +"\""+ path_to_file +"\""+ '>>' +"\""+ help_dir_name + file_name + '-copy.txt'+"\""
			print('\nВстретился .doc файл....')
			os.system(sys_query)

			for cell in sheet['A']:
				pattern = cell.value.strip()
				with open(help_dir_name + file_name + '-copy.txt','r') as subFile:
					ins = 0
					for line in subFile:
						ins += find_key(pattern, line)
					if ins > 5:
						print('Успешно!')
						if matches.get(pattern) != None:
							matches[pattern] += [path_to_file]
						else:
							matches[pattern] = [path_to_file]
						
						i += 1
						bar.update(i)
		except:

			print('\n Данный файл ' + path_to_file + ' вызвал ошибку', file=errors)
			
				

errors.close()

with open(desktop + 'proofs.txt','w') as output:
	print("№     КЛЮЧЕВОЕ СЛОВО                         ФАЙЛЫ", file=output)
	for cell in sheet['A']:
		if matches.get(cell.value.strip()) != None:
			print("\n"+str(cell.row)+"   | "+cell.value.strip()+" | ->  ", file=output, end='')
			for vals in matches[cell.value.strip()]:
				print ('    '+ vals, file=output, end='')
		else:
			print("\n"+str(cell.row)+"   | "+cell.value.strip()+" | ------  0 ", file=output, end='')

j = 0
with open(desktop + 'matches.txt','w') as output:
	for key in matches:
		print("\n По этому запросу - '"+key+"' нашлись следующие файлы: ", file=output)
		for vals in matches[key]:
			j += 1
			print ('    '+str(j)+'. '+vals, file=output)
		j = 0




response = ''
while(response != 'y'):
	response = input('Пожалуйста, просмотрите файл Matches.txt на вашем рабочем столе\nЗапустить скрипт удаления файлов?\n\n[Y]').lower()
	if response == 'y':
		system_query = r'python delete.py'
		os.system(system_query)