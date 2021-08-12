# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from getAllinCol import get_all_in_col_by_title
from dollar import get_current_exchange
code = 'R01235' #код доллара ЦБ
wb = load_workbook('list.xlsx')
sheet = wb[wb.sheetnames[0]]



mass = get_all_in_col_by_title(sheet, "ID")
print (get_current_exchange(code))
